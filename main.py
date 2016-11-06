import requests
import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Alignment
from requests_ntlm import HttpNtlmAuth

url = "http://url/dummy/"
username = input("User: ")
username = "DOMAIN\\" + username
password = input("Pass: ")

# html = requests.get(url, auth=HttpNtlmAuth(username=username, password=password))
html = open('dummy.html', 'r')
soup = BeautifulSoup(html, "html5lib")

projects = soup.find_all('div', {'class': 'ParentRow'})
projectsAlt = soup.find_all('div', {'class': 'ParentRowAlt'})

projectInfo = []
projectInfoAlt = []
projectsInfo = [[]]

# Function that takes list of project info and writes it into excel file
def writeToExcel(projectsList, file):
    wb = load_workbook(file)
    # Selects the active worksheet
    ws = wb.active

    # Adds company logo
    # img = Image('logo.png')
    # ws.add_image(img, 'D1')

    # Removes the last record (project) form the list as it appears to be newline
    projectsList.pop()

    r, c = 5, 1
    for p in projectsList:
        for data in p:
            # Converts the first record of each project (project number) to integer
            if c == 1:
                data = int(data)

            ws.cell(row=r, column=c).value = data
            c += 2

        # Loop that goes trough all the columns and colors each cell
        for counter in range(1, 10):
            fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            ws.cell(row=r, column=counter).fill = fill
            # ws.cell(row=r, column=counter).alignment.vertical = 'top'
            # if counter == 2:
            #     ws.cell(row=r, column=counter).alignment = Alignment(wrapText=True)

        r += 1
        c = 1

    # Get the current datetime
    time = datetime.datetime.now()
    # Format the datetime and convert it to string
    time = time.strftime('%m-%d-%Y')

    wb.save('Level 1 Project - ' + time + '.xlsx')


# Function that takes project html and return list with Project number, Start Date, End Date and Project Level
def getProjectInfo(p):
    projectNum = p.find('div', {'class': 'ProjectNoCol'})
    projectNum = projectNum.a.text

    projectFacility = p.find('div', {'class': 'FacilityName'})
    projectFacility = projectFacility.text

    projectStartTime = p.find('div', {'class': 'DateCol'})
    projectStartTime = projectStartTime.text

    projectEndTime = p.find('div', {'class': 'DateColTo'})
    projectEndTime = projectEndTime.text

    projectLevel = p.find('div', {'class': 'TwoChanel'})
    projectLevel = projectLevel.text

    # Removing the '+' from the project number
    if len(projectNum) == 8:
        projectNum = projectNum[:-1]
    projectNum = projectNum.strip()

    # Removing label from facility name
    projectFacility = projectFacility[10:]
    projectFacility = projectFacility.strip()

    # Remove the empty spaces and the label from the start time
    projectStartTime = projectStartTime.replace(' ', '')
    projectStartTime = projectStartTime[10:]

    # Remove the empty spaces and the label from the end time
    projectEndTime = projectEndTime.replace(' ', '')
    projectEndTime = projectEndTime[3:]

    # Fills the project info into list
    pInfo = [projectNum, projectFacility, projectStartTime, projectEndTime, projectLevel]

    return pInfo

# The row in the 2D list that for each project
row = 0
# Goes through list with project's html
for project in projects:
    level = project.find('div', {'class': 'TwoChanel'})

    # Checks if the project is Level 1
    if level.text != '':
        projectsInfo.append([])
        projectInfo = getProjectInfo(project)

        for info in projectInfo:
            projectsInfo[row].append(info)
        row += 1

# Goes through list with projectAlt htmls
for projectAlt in projectsAlt:
    level = projectAlt.find('div', {'class': 'TwoChanel'})

    # Checks if the projectAlt is Level 1
    if level.text != '':
        projectsInfo.append([])
        projectInfoAlt = getProjectInfo(projectAlt)

        for info in projectInfoAlt:
            projectsInfo[row].append(info)
        row += 1

# Loop that goes through the 2D lists with all the projects info and prints each project as a string
for project in projectsInfo:
    dataString = ' | '.join(project)
    print(dataString)


# writeToExcel(projectsInfo, 'template.xlsx')
