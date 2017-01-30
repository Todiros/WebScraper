from __future__ import print_function
import datetime
import warnings
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from selenium import webdriver

def main():
  warnings.filterwarnings('ignore')
  
  driver = webdriver.Firefox()
  driver.get("http://private-url/projects/")
  html = driver.page_source
  soup = BeautifulSoup(html, "html5lib")
  driver.quit()
  
  projects = soup.find_all('div', {'class': 'ParentRow'})
  projectsAlt = soup.find_all('div', {'class': 'ParentRowAlt'})

  projectsHtml = projects + projectsAlt
  projectData = []		
  
  projectsData = [[]]
  
  # Function that takes list of projects info and writes it into excel file
  def writeToExcel(projectsList, file):
      wb = load_workbook(file)
  
      # Selects the active worksheet
      ws = wb.active
  
      # Adds company logo
      img = Image('logo.png')
      ws.add_image(img, 'D1')
  
      #starting cell coordinates (5th row, 1st column)	
      r, c = 5, 1

      for p in projectsList:
          for data in p:
              # Converts the first record of each project (project number) to integer
              if c == 1:
                  data = int(data)
  
              ws.cell(row=r, column=c).value = data
              c += 2
  
          # Loop that goes trough all the columns and colors each cell
          for counter in range(1, 10):
              fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
              ws.cell(row=r, column=counter).fill = fill
  
          r += 1
          c = 1
  
      # Get the current datetime
      date = datetime.datetime.now()
  
      # Format the datetime and convert it to string
      date = date.strftime('%m-%d-%Y')
  
      wb.save('Level 1 Projects - ' + date + '.xlsx')
  
  
  # Function that takes project html and return list with Project number, Start Date, End Date and Project Level
  def getProjectData(p):
      projectNum = p.find('div', {'class': 'ProjectNoCol'})
      projectNum = projectNum.a.text
  
      projectFacility = p.find('div', {'class': 'FacilityName'})
      projectFacility = projectFacility.text
  
      projectStartTime = p.find('div', {'class': 'DateCol'})
      projectStartTime = projectStartTime.text
  
      projectEndTime = p.find('div', {'class': 'DateColTo'})
      projectEndTime = projectEndTime.text
  
      projectLevel = p.find('div', {'class': 'TwoChanel'})
      projectLevel = projectLevel.text
  
      # Removing the '+' from the project number
      projectNum = ''.join(projectNum.split())
      if len(projectNum) >= 7:	
      	projectNum = projectNum[:-1]
  
      # Removing label from facility name
      projectFacility = projectFacility[10:]
      projectFacility = projectFacility.strip()
  
      # Remove the empty spaces and the label from the start time
      projectStartTime = projectStartTime.replace(' ', '')
      projectStartTime = projectStartTime[10:]
  
      # Remove the empty spaces and the label from the end time
      projectEndTime = projectEndTime.replace(' ', '')
      projectEndTime = projectEndTime[3:]
  
      # Fills the project info into list
      pInfo = [projectNum, projectFacility, projectStartTime, projectEndTime, projectLevel]
  
      return pInfo
  
  # The row in the 2D list for each project
  row = 0
  
  # Goes through list with project's html
  for project in projectsHtml:
    level = project.find('div', {'class': 'TwoChanel'})

    # Checks if the project is Level 1
    if level.text != '':
        projectsData.append([])
        projectData = getProjectData(project)

        for info in projectData:
            projectsData[row].append(info)
        row += 1

  #Checks if any Level 1 Projects had been found
  if not (projectData):
       print("No Level 1 Projects for today!")
       quit()	
  
  # Removes the last record (project) form the list as it appears to be newline
  projectsData.pop()
  
  # Sorts the list by project number
  projectsData = sorted(projectsData,key=lambda l:l[0])
  
  # Loop that goes through the 2D lists with all the projects info and prints it as a string
  print('\n')

  for project in projectsData:
      dataString = ' | '.join(project)
      print(dataString)

  print('--------------------')
  print('Level 1 Projects: ', len(projectsData), '\n')
  
  writeToExcel(projectsData, 'template.xlsx')

if __name__ == "__main__":
    main()
